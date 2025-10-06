import os
import asyncpg
from typing import Optional, Any, Dict
import logging

logger = logging.getLogger("TusaBot")

# Environment with sane defaults based on your provided DB creds
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME", "FamilyDB")
DB_USER = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "1")


async def create_pool() -> asyncpg.Pool:
    return await asyncpg.create_pool(
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        min_size=1,
        max_size=10,
        command_timeout=30,
    )


async def init_schema(pool: asyncpg.Pool) -> None:
    async with pool.acquire() as conn:
        # Таблица пользователей
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                tg_id BIGINT PRIMARY KEY,
                name TEXT,
                gender TEXT CHECK (gender IN ('male', 'female')),
                age INTEGER CHECK (age >= 16 AND age <= 100),
                vk_id TEXT,
                username TEXT,
                registered_at TIMESTAMPTZ DEFAULT now(),
                created_at TIMESTAMPTZ DEFAULT now(),
                updated_at TIMESTAMPTZ DEFAULT now()
            );
            """
        )
        
        # Таблица афиш
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS posters (
                id SERIAL PRIMARY KEY,
                file_id TEXT NOT NULL,
                caption TEXT,
                ticket_url TEXT,
                created_at TIMESTAMPTZ DEFAULT now(),
                is_active BOOLEAN DEFAULT true
            );
            """
        )
        
        # Таблица посещаемости
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS attendances (
                id SERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(tg_id) ON DELETE CASCADE,
                poster_id INTEGER REFERENCES posters(id) ON DELETE CASCADE,
                attended_at TIMESTAMPTZ DEFAULT now(),
                UNIQUE(user_id, poster_id)
            );
            """
        )
        
        # Индексы
        await conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_users_vk_id ON users(vk_id);
            CREATE INDEX IF NOT EXISTS idx_users_registered_at ON users(registered_at);
            CREATE INDEX IF NOT EXISTS idx_posters_is_active ON posters(is_active);
            CREATE INDEX IF NOT EXISTS idx_attendances_user_id ON attendances(user_id);
            CREATE INDEX IF NOT EXISTS idx_attendances_poster_id ON attendances(poster_id);
            """
        )
        
        # Создание функции для автоматического обновления updated_at
        await conn.execute(
            """
            CREATE OR REPLACE FUNCTION update_updated_at_column()
            RETURNS TRIGGER AS $$
            BEGIN
                NEW.updated_at = now();
                RETURN NEW;
            END;
            $$ language 'plpgsql';
            """
        )
        
        # Создание триггера для автоматического обновления updated_at в таблице users
        await conn.execute(
            """
            DROP TRIGGER IF EXISTS update_users_updated_at ON users;
            """
        )
        await conn.execute(
            """
            CREATE TRIGGER update_users_updated_at 
                BEFORE UPDATE ON users 
                FOR EACH ROW 
                EXECUTE FUNCTION update_updated_at_column();
            """
        )


async def upsert_user(
    pool: asyncpg.Pool,
    tg_id: int,
    name: Optional[str] = None,
    gender: Optional[str] = None,
    age: Optional[int] = None,
    vk_id: Optional[str] = None,
    username: Optional[str] = None,
) -> None:
    async with pool.acquire() as conn:
        try:
            logger.info("Upserting user %s: name=%s, gender=%s, age=%s, vk_id=%s, username=%s", 
                       tg_id, name, gender, age, vk_id, username)
            await conn.execute(
                """
                INSERT INTO users (tg_id, name, gender, age, vk_id, username)
                VALUES ($1, $2, $3, $4, $5, $6)
                ON CONFLICT (tg_id) DO UPDATE
                SET name = COALESCE(EXCLUDED.name, users.name),
                    gender = COALESCE(EXCLUDED.gender, users.gender),
                    age = COALESCE(EXCLUDED.age, users.age),
                    vk_id = COALESCE(EXCLUDED.vk_id, users.vk_id),
                    username = COALESCE(EXCLUDED.username, users.username);
                """,
                tg_id,
                name,
                gender,
                age,
                vk_id,
                username,
            )
            logger.info("Successfully upserted user %s", tg_id)
        except Exception as e:
            logger.error("Failed to upsert user %s: %s", tg_id, e)
            raise


async def set_vk_id(pool: asyncpg.Pool, tg_id: int, vk_id: str) -> None:
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE users SET vk_id=$2 WHERE tg_id=$1",
            tg_id,
            vk_id,
        )


async def get_user(pool: asyncpg.Pool, tg_id: int) -> Optional[Dict[str, Any]]:
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM users WHERE tg_id=$1", tg_id)
        return dict(row) if row else None


async def get_user_by_username(pool: asyncpg.Pool, username: str) -> Optional[Dict[str, Any]]:
    """Поиск пользователя по Telegram username"""
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM users WHERE LOWER(username)=LOWER($1)", username)
        return dict(row) if row else None


async def get_all_user_ids(pool: asyncpg.Pool) -> list[int]:
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT tg_id FROM users")
        return [r[0] for r in rows]


async def load_user_vk_data(pool: asyncpg.Pool) -> dict[int, str]:
    """Загрузить VK ID всех пользователей для кеширования"""
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT tg_id, vk_id FROM users WHERE vk_id IS NOT NULL")
        return {row[0]: row[1] for row in rows}


async def get_user_stats(pool: asyncpg.Pool) -> dict:
    """Получить статистику пользователей"""
    async with pool.acquire() as conn:
        stats = await conn.fetchrow("""
            SELECT 
                COUNT(*) as total_users,
                COUNT(vk_id) as users_with_vk,
                COUNT(CASE WHEN gender = 'male' THEN 1 END) as male_users,
                COUNT(CASE WHEN gender = 'female' THEN 1 END) as female_users,
                COUNT(CASE WHEN registered_at >= CURRENT_DATE THEN 1 END) as today_registrations
            FROM users
        """)
        return dict(stats) if stats else {}


async def export_users_to_excel(pool: asyncpg.Pool, filename: str = "users_export.xlsx") -> str:
    """Экспорт всех пользователей в Excel файл"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        async with pool.acquire() as conn:
            users = await conn.fetch("""
                SELECT 
                    tg_id,
                    name,
                    gender,
                    age,
                    vk_id,
                    registered_at,
                    created_at
                FROM users 
                ORDER BY registered_at DESC
            """)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи TusaBot"
        
        # Заголовки
        headers = [
            "Telegram ID", "Имя", "Пол", "Возраст", 
            "VK ID", "Дата регистрации", "Дата создания"
        ]
        
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные пользователей
        for row_idx, user in enumerate(users, 2):
            ws.cell(row=row_idx, column=1, value=user['tg_id'])
            ws.cell(row=row_idx, column=2, value=user['name'] or "Не указано")
            
            # Пол на русском
            gender_map = {"male": "Мужской", "female": "Женский"}
            ws.cell(row=row_idx, column=3, value=gender_map.get(user['gender'], "Не указано"))
            
            # Возраст
            age = user['age']
            if age:
                ws.cell(row=row_idx, column=4, value=f"{age} лет")
            else:
                ws.cell(row=row_idx, column=4, value="Не указано")
            
            ws.cell(row=row_idx, column=5, value=user['vk_id'] or "Не привязан")
            
            # Даты
            if user['registered_at']:
                ws.cell(row=row_idx, column=6, value=user['registered_at'].strftime("%d.%m.%Y %H:%M"))
            else:
                ws.cell(row=row_idx, column=6, value="Не указано")
                
            if user['created_at']:
                ws.cell(row=row_idx, column=7, value=user['created_at'].strftime("%d.%m.%Y %H:%M"))
            else:
                ws.cell(row=row_idx, column=7, value="Не указано")
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем лист со статистикой
        stats_ws = wb.create_sheet("Статистика")
        stats = await get_user_stats(pool)
        
        stats_data = [
            ["Показатель", "Значение"],
            ["Всего пользователей", stats.get('total_users', 0)],
            ["С привязанным VK", stats.get('users_with_vk', 0)],
            ["Мужчин", stats.get('male_users', 0)],
            ["Женщин", stats.get('female_users', 0)],
            ["Зарегистрировано сегодня", stats.get('today_registrations', 0)],
            ["Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M")]
        ]
        
        for row_idx, (label, value) in enumerate(stats_data, 1):
            stats_ws.cell(row=row_idx, column=1, value=label)
            stats_ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:  # Заголовок
                stats_ws.cell(row=row_idx, column=1).font = header_font
                stats_ws.cell(row=row_idx, column=2).font = header_font
                stats_ws.cell(row=row_idx, column=1).fill = header_fill
                stats_ws.cell(row=row_idx, column=2).fill = header_fill
        
        # Автоподбор ширины для статистики
        for column in stats_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            stats_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(filename)
        return filename
        
    except Exception as e:
        raise Exception(f"Ошибка экспорта в Excel: {e}")


# ----------------------
# Функции для работы с афишами (posters)
# ----------------------

async def create_poster(
    pool: asyncpg.Pool,
    file_id: str,
    caption: Optional[str] = None,
    ticket_url: Optional[str] = None,
) -> int:
    """Создать новую афишу и вернуть её ID"""
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO posters (file_id, caption, ticket_url, is_active)
            VALUES ($1, $2, $3, true)
            RETURNING id
            """,
            file_id,
            caption,
            ticket_url,
        )
        return row['id']


async def get_active_posters(pool: asyncpg.Pool) -> list[Dict[str, Any]]:
    """Получить все активные афиши"""
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, file_id, caption, ticket_url, created_at, is_active
            FROM posters
            WHERE is_active = true
            ORDER BY created_at DESC
            """
        )
        return [dict(row) for row in rows]


async def get_latest_poster(pool: asyncpg.Pool) -> Optional[Dict[str, Any]]:
    """Получить последнюю активную афишу"""
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, file_id, caption, ticket_url, created_at, is_active
            FROM posters
            WHERE is_active = true
            ORDER BY created_at DESC
            LIMIT 1
            """
        )
        return dict(row) if row else None


async def get_poster_by_id(pool: asyncpg.Pool, poster_id: int) -> Optional[Dict[str, Any]]:
    """Получить афишу по ID"""
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, file_id, caption, ticket_url, created_at, is_active FROM posters WHERE id=$1",
            poster_id
        )
        return dict(row) if row else None


async def deactivate_poster(pool: asyncpg.Pool, poster_id: int) -> None:
    """Деактивировать афишу (мягкое удаление)"""
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE posters SET is_active=false WHERE id=$1",
            poster_id
        )


async def delete_poster(pool: asyncpg.Pool, poster_id: int) -> None:
    """Удалить афишу полностью"""
    async with pool.acquire() as conn:
        await conn.execute(
            "DELETE FROM posters WHERE id=$1",
            poster_id
        )


async def update_poster_ticket_url(pool: asyncpg.Pool, poster_id: int, ticket_url: str) -> None:
    """Обновить ссылку на билеты для афиши"""
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE posters SET ticket_url=$2 WHERE id=$1",
            poster_id,
            ticket_url
        )


# ----------------------
# Функции для работы с посещаемостью (attendances)
# ----------------------

async def mark_attendance(pool: asyncpg.Pool, user_id: int, poster_id: int) -> bool:
    """Отметить посещение пользователя на мероприятии. Возвращает True если успешно, False если уже отмечен"""
    async with pool.acquire() as conn:
        try:
            await conn.execute(
                """
                INSERT INTO attendances (user_id, poster_id)
                VALUES ($1, $2)
                ON CONFLICT (user_id, poster_id) DO NOTHING
                """,
                user_id,
                poster_id
            )
            return True
        except Exception:
            return False


async def get_user_attendances(pool: asyncpg.Pool, user_id: int) -> list[Dict[str, Any]]:
    """Получить все посещения пользователя"""
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT a.id, a.poster_id, a.attended_at, p.caption
            FROM attendances a
            JOIN posters p ON a.poster_id = p.id
            WHERE a.user_id = $1
            ORDER BY a.attended_at DESC
            """,
            user_id
        )
        return [dict(row) for row in rows]


async def get_poster_attendances(pool: asyncpg.Pool, poster_id: int) -> list[Dict[str, Any]]:
    """Получить всех пользователей, посетивших мероприятие"""
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT a.id, a.user_id, a.attended_at, u.name, u.gender, u.age
            FROM attendances a
            JOIN users u ON a.user_id = u.tg_id
            WHERE a.poster_id = $1
            ORDER BY a.attended_at DESC
            """,
            poster_id
        )
        return [dict(row) for row in rows]


async def get_attendance_stats(pool: asyncpg.Pool, poster_id: int) -> dict:
    """Получить статистику посещаемости мероприятия"""
    async with pool.acquire() as conn:
        stats = await conn.fetchrow(
            """
            SELECT 
                COUNT(*) as total_attendees,
                COUNT(CASE WHEN u.gender = 'male' THEN 1 END) as male_count,
                COUNT(CASE WHEN u.gender = 'female' THEN 1 END) as female_count
            FROM attendances a
            JOIN users u ON a.user_id = u.tg_id
            WHERE a.poster_id = $1
            """,
            poster_id
        )
        return dict(stats) if stats else {}
